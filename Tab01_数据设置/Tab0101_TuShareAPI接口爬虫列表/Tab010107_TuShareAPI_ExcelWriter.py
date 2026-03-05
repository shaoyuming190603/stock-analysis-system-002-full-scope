# This Python file uses the following encoding: utf-8

# if __name__ == "__main__":
#     pass
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from utils import ensure_dir

class ExcelWriter:
    def __init__(self, save_dir, excel_name, logger):
        ensure_dir(save_dir)
        self.file_path = os.path.join(save_dir, excel_name)
        self.logger = logger

    def write(self, all_data):
        wb = Workbook()
        # 六个sheet
        sheets = [
            ("输入参数_名称", "input_params"),
            ("输入参数_类型", "input_types"),
            ("输入参数_默认显示", "input_defaults"),
            ("输出参数_名称", "output_params"),
            ("输出参数_类型", "output_types"),
            ("输出参数_默认显示", "output_defaults"),
        ]
        for idx, (sheet_name, key) in enumerate(sheets):
            ws = wb.create_sheet(title=sheet_name, index=idx)
            max_len = max((len(d[key]) for d in all_data), default=0)
            header = ["接口", "标题", f"{'输入' if idx < 3 else '输出'}参数数"] + [f"参数{i+1}" for i in range(max_len)]
            ws.append(header)
            for d in all_data:
                row = [d["api_name"], d["title"], len(d[key])] + d[key] + [""] * (max_len - len(d[key]))
                ws.append(row)
                # 标题加超链接
                if ws.max_row == 2:
                    ws.cell(row=2, column=2).hyperlink = d["url"]
                    ws.cell(row=2, column=2).font = Font(color="0000FF", underline="single")
        # 删除默认sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(self.file_path)
        self.logger.log(f"Excel已保存：{self.file_path}")
