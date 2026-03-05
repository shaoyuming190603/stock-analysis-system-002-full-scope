import os
import requests
from bs4 import BeautifulSoup
import time
import re
import logging
from datetime import datetime
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

class TushareDocSpider:
    """
    Tushare Pro 文档爬虫（增强版）：提取接口名称、页面大标题、输入/输出参数的详细信息
    （名称、类型、显示、描述），并生成包含八个工作表的 Excel 文件。
    """
    def __init__(self, start_url, output_file='tushare_api_docs.xlsx', log_file='spider.log', delay=1):
        self.start_url = start_url
        self.output_file = output_file
        self.delay = delay
        self.log_file = log_file

        # 配置日志记录器
        self._setup_logging()

        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })

        # 存储结构：每个元素为 (接口名, 页面大标题, 页面URL, 参数详细信息列表)
        # 参数详细信息为字典，包含 'name', 'type', 'display', 'desc' 键
        self.input_data = [] # 输入参数详细信息
        self.output_data = [] # 输出参数详细信息

        # 统计信息
        self.total_pages = 0
        self.success_pages = 0
        self.skipped_pages = 0
        self.failed_pages = 0

    def _setup_logging(self):
        """配置日志记录器，同时输出到控制台和文件"""
        self.logger = logging.getLogger('TushareSpider')
        self.logger.setLevel(logging.INFO)

        if self.logger.handlers:
            self.logger.handlers.clear()

        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

        log_dir = os.path.dirname(self.log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir)
        file_handler = logging.FileHandler(self.log_file, encoding='utf-8')
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)

        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        self.logger.addHandler(console_handler)

    def get_all_doc_links(self):
        """从起始页面提取所有接口文档的链接（包含 doc_id 的）"""
        self.logger.info(f"正在从起始页 {self.start_url} 提取所有文档链接...")
        try:
            resp = self.session.get(self.start_url)
            resp.encoding = 'utf-8'
            soup = BeautifulSoup(resp.text, 'html.parser')
            links = set()
            for a in soup.find_all('a', href=True):
                href = a['href']
                if 'document/2?doc_id=' in href:
                    full_url = urljoin(self.start_url, href)
                    links.add(full_url)
            self.logger.info(f"共找到 {len(links)} 个接口文档链接。")
            return list(links)
        except Exception as e:
            self.logger.error(f"获取链接失败：{e}", exc_info=True)
            return []

    def extract_title_above_interface(self, soup):
        """
        提取页面的大标题。
        改进点：
        1. 优先直接查找页面中的 h1/h2 标签，并过滤掉小节标题（如“输入参数”等），
           以覆盖特殊页面（如 doc_id=367/109/366）中标题位于 h2 的情况。
        2. 如果上述方法未找到合适的标题，则回退到原逻辑：查找“接口：”上方的标题。
        """
        # 常见的小节标题关键词，这些不应作为页面大标题
        section_keywords = ['输入参数', '输出参数', '返回数据', '错误码', '参数说明', '注意事项', '更新记录']

        # 1. 尝试找 h1 标签
        h1 = soup.find('h1')
        if h1:
            title = h1.get_text().strip()
            if title and not any(keyword in title for keyword in section_keywords):
                return title

        # 2. 尝试找 h2 标签
        h2 = soup.find('h2')
        if h2:
            title = h2.get_text().strip()
            if title and not any(keyword in title for keyword in section_keywords):
                return title

        # 3. 回退到原逻辑：查找“接口：”上方的标题
        interface_str = soup.find(string=re.compile(r'接口[：:]'))
        if not interface_str:
            # 如果没有“接口：”，尝试查找“接口名称：”
            interface_str = soup.find(string=re.compile(r'接口名称[：:]'))
        if not interface_str:
            return ""

        tag = interface_str.parent
        title_tag = tag.find_previous(['h1', 'h2', 'h3'])
        if not title_tag:
            title_tag = tag.find_previous(class_=lambda x: x and 'title' in x.lower())

        if title_tag:
            return title_tag.get_text().strip()
        return ""

    def _extract_params_details(self, soup, table_title):
        """
        从页面中提取指定标题的表格中的参数详细信息。
        返回一个列表，每个元素是一个字典，包含 'name', 'type', 'display', 'desc' 键。
        """
        details = []
        # 找到包含表格标题的标签
        header = soup.find(lambda tag: tag.name in ['h3', 'h4', 'div', 'p'] and table_title in tag.get_text())
        if not header:
            return details
        table = header.find_next('table')
        if not table:
            return details

        rows = table.find_all('tr')
        if len(rows) < 2:
            return details

        # 尝试识别表头行（第一行）
        header_row = rows[0]
        header_cells = header_row.find_all(['th', 'td'])
        header_texts = [cell.get_text().strip().lower() for cell in header_cells]

        # 定义字段关键词映射
        field_keywords = {
            'name': ['名称', '参数名', '字段', 'name'],
            'type': ['类型', 'type'],
            'desc': ['描述', '说明', 'desc'],
            'display': ['默认', '显示', '示例', 'default', 'display']
        }

        # 建立列索引到字段名的映射
        col_to_field = {}
        for idx, text in enumerate(header_texts):
            for field, keywords in field_keywords.items():
                if any(keyword in text for keyword in keywords):
                    col_to_field[idx] = field
                    break

        # 如果没有识别到任何字段，则使用默认顺序：第一列name，第二列type，第三列desc，第四列display
        if not col_to_field:
            for idx in range(len(header_cells)):
                if idx == 0:
                    col_to_field[idx] = 'name'
                elif idx == 1:
                    col_to_field[idx] = 'type'
                elif idx == 2:
                    col_to_field[idx] = 'desc'
                elif idx == 3:
                    col_to_field[idx] = 'display'
                # 多余列忽略

        # 遍历数据行（跳过表头行）
        for row in rows[1:]:
            cells = row.find_all('td')
            if not cells:
                continue
            param = {'name': '', 'type': '', 'desc': '', 'display': ''}
            for idx, cell in enumerate(cells):
                if idx in col_to_field:
                    field = col_to_field[idx]
                    param[field] = cell.get_text().strip()
            if param['name']: # 只保留有名称的参数
                details.append(param)
        return details

    def parse_api_page(self, url):
        """
        解析单个接口页面，提取所有需要的信息
        返回： (api_name, title_above, url, input_details, output_details)
        如果接口名称为 'index_classify'，则返回 (None, None, None, None, None) 表示跳过
        """
        self.logger.info(f"正在解析：{url}")
        try:
            resp = self.session.get(url)
            resp.encoding = 'utf-8'
            soup = BeautifulSoup(resp.text, 'html.parser')
            text = soup.get_text()

            # 1. 提取接口名称（增强对“接口名称：”的识别）
            api_name = None
            # 先尝试匹配“接口：”或“接口名称：”
            interface_pattern = re.compile(r'(?:接口|接口名称)[：:]\s*([a-zA-Z0-9_]+)')
            match = interface_pattern.search(text)
            if match:
                api_name = match.group(1)
            else:
                # 如果正则匹配不到，则遍历标签
                for tag in soup.find_all(['h1', 'h2', 'h3', 'p', 'strong']):
                    if '接口' in tag.get_text():
                        text_line = tag.get_text()
                        parts = re.split(r'[：:]', text_line)
                        if len(parts) > 1:
                            candidate = parts[1].strip().split()[0]
                            if candidate and re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', candidate):
                                api_name = candidate
                                break

            if not api_name:
                self.logger.warning(f"无法从 {url} 提取接口名称，跳过。")
                self.failed_pages += 1
                return None, None, None, None, None

            # 检查是否需要跳过
            if api_name == 'index_classify':
                self.logger.info(f"接口名称为 '{api_name}'，根据要求跳过此页面。")
                self.skipped_pages += 1
                return None, None, None, None, None

            # 2. 提取页面大标题（使用改进后的方法）
            title_above = self.extract_title_above_interface(soup)
            if not title_above:
                self.logger.warning(f"页面 {url} 未找到合适的标题，将使用空字符串。")

            # 3. 提取输入参数详细信息
            input_details = self._extract_params_details(soup, '输入参数')
            # 4. 提取输出参数详细信息
            output_details = self._extract_params_details(soup, '输出参数')

            self.logger.info(f"成功解析接口：{api_name}，标题：{title_above}，输入参数数：{len(input_details)}，输出参数数：{len(output_details)}")
            self.success_pages += 1
            return api_name, title_above, url, input_details, output_details

        except Exception as e:
            self.logger.error(f"解析页面 {url} 时出错：{e}", exc_info=True)
            self.failed_pages += 1
            return None, None, None, None, None

    def run(self):
        """执行爬虫主流程"""
        self.logger.info("="*50)
        self.logger.info("爬虫程序开始运行")
        self.logger.info(f"目标起始页：{self.start_url}")
        self.logger.info(f"输出文件：{self.output_file}")
        self.logger.info(f"日志文件：{self.log_file}")
        self.logger.info("="*50)

        # 1. 获取所有链接
        doc_links = self.get_all_doc_links()
        if not doc_links:
            self.logger.error("未找到任何文档链接，程序退出。")
            return

        self.total_pages = len(doc_links)
        self.logger.info(f"开始遍历 {self.total_pages} 个接口页面...")

        # 2. 遍历每个链接，解析数据
        for idx, link in enumerate(doc_links):
            self.logger.info(f"进度：{idx+1}/{self.total_pages}")
            result = self.parse_api_page(link)
            if result[0]: # 只有成功解析且非跳过的才会被记录
                api_name, title_above, url, input_details, output_details = result
                self.input_data.append((api_name, title_above, url, input_details))
                self.output_data.append((api_name, title_above, url, output_details))
            time.sleep(self.delay)

        # 3. 输出统计信息
        self.logger.info("="*50)
        self.logger.info("爬取完成，统计信息：")
        self.logger.info(f"总页面数：{self.total_pages}")
        self.logger.info(f"成功解析：{self.success_pages}")
        self.logger.info(f"跳过页面：{self.skipped_pages}")
        self.logger.info(f"失败页面：{self.failed_pages}")
        self.logger.info("="*50)

        # 4. 输出到Excel
        self._save_to_excel()

    def _apply_styles(self, worksheet):
        """
        为工作表的整个数据区（包括表头）应用点线边框和 Arial 11 号字体。
        超链接单元格保留蓝色下划线。
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row < 1 or max_col < 1:
            return

        dotted_side = Side(style='dotted')

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)

                # 设置边框：所有单元格都有右边框和下边框，第一列有左边框，第一行有上边框
                left = dotted_side if c == 1 else None
                top = dotted_side if r == 1 else None
                right = dotted_side # 所有列都有右边框
                bottom = dotted_side # 所有行都有下边框
                border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.border = border

                # 设置字体：普通单元格 Arial 11，超链接单元格保留蓝色下划线
                if cell.hyperlink:
                    font = Font(name='Arial', size=11, color="0000FF", underline="single")
                else:
                    font = Font(name='Arial', size=11)
                cell.font = font

    def _save_to_excel(self):
        """将数据保存为Excel文件，包含八个工作表，第二列根据需要设置超链接，并应用样式和标签颜色"""
        self.logger.info("开始生成Excel文件...")

        # 确保输出目录存在
        output_dir = os.path.dirname(self.output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.logger.info(f"已创建目录：{output_dir}")

        # 定义固定的列名：接口、描述、参数数 + 20个参数列
        max_params = 20
        columns = ['接口', '描述', '参数数'] + [f'参数{i+1}' for i in range(max_params)]

        # 工作表配置： (标题, 数据源, 属性键, 是否超链接)
        sheets_config = [
            ('输入参数', self.input_data, 'name', True),
            ('输入类型', self.input_data, 'type', False),
            ('输入显示', self.input_data, 'display', False),
            ('输入描述', self.input_data, 'desc', False),
            ('输出参数', self.output_data, 'name', True),
            ('输出类型', self.output_data, 'type', False),
            ('输出显示', self.output_data, 'display', False),
            ('输出描述', self.output_data, 'desc', False)
        ]

        # 工作表标签颜色（VBA 颜色值转换为 RGB 十六进制）
        tab_colors = {
            '输入参数': 'FF0000', # 红色
            '输入类型': '9933FF', # 紫色 (10040319)
            '输入显示': '9933FF',
            '输入描述': '9933FF',
            '输出参数': '0000FF', # 蓝色 (16711680)
            '输出类型': '00FF00', # 绿色（近似主题色 Accent5 + 亮度）
            '输出显示': '00FF00',
            '输出描述': '00FF00'
        }

        # 创建Excel工作簿
        wb = Workbook()
        # 移除默认创建的Sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # 辅助函数：为每个工作表准备行数据
        def prepare_rows(data, attr_key, with_hyperlink):
            """
            data: 列表，每个元素为 (api, title, url, details)
            attr_key: 要从details中提取的属性键
            with_hyperlink: 第二列是否为超链接
            返回行列表，每行为一个列表，其中第二列可能是字符串或(文本, url)元组
            """
            rows = []
            for api, title, url, details in data:
                # 提取所有参数的指定属性值
                values = [d.get(attr_key, '') for d in details]
                param_count = len(values)
                # 填充到最大列数
                values_padded = values[:max_params] + [''] * (max_params - len(values))
                # 构建行
                if with_hyperlink:
                    row = [api, (title, url), param_count] + values_padded
                else:
                    row = [api, title, param_count] + values_padded
                rows.append(row)
            return rows

        # 循环创建每个工作表
        for title, data_source, attr_key, hyperlink in sheets_config:
            ws = wb.create_sheet(title=title)
            ws.append(columns)

            rows = prepare_rows(data_source, attr_key, hyperlink)
            for row_data in rows:
                # 构建要写入的行（将第二列的元组拆分为文本）
                row_to_write = []
                for cell in row_data:
                    if isinstance(cell, tuple) and len(cell) == 2:
                        row_to_write.append(cell[0]) # 只添加显示文本
                    else:
                        row_to_write.append(cell)
                ws.append(row_to_write)

                # 如果需要超链接，设置第二列的 hyperlink 属性（字体稍后统一设置）
                if hyperlink:
                    last_row = ws.max_row
                    cell = ws.cell(row=last_row, column=2)
                    cell.hyperlink = row_data[1][1] # 超链接地址

            # 调整列宽
            ws.column_dimensions['A'].width = 20 # 接口列
            ws.column_dimensions['B'].width = 40 # 描述列
            ws.column_dimensions['C'].width = 15 # 参数数列
            for i in range(4, 4 + max_params): # 参数列从D列开始
                ws.column_dimensions[get_column_letter(i)].width = 15

            # 应用边框和字体样式
            self._apply_styles(ws)

            # 设置工作表标签颜色
            ws.sheet_properties.tabColor = tab_colors[title]

        wb.save(self.output_file)
        self.logger.info(f"Excel文件已成功保存到：{self.output_file}")

if __name__ == "__main__":
    # 指定保存路径和文件名
    base_dir = r"C:\01_My_documents\01_pythonLearn\我的PNP_MCSD大项目\Tab01_数据设置\Tab0101_TuShareAPI接口参数"
    save_filename = "tuShare接口输入输出参数清单Ae9.xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"tuShare爬虫日志_{timestamp}.log"

    logging.basicConfig(filename=log_filename, level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s')

    full_excel_path = os.path.join(base_dir, save_filename)
    full_log_path = os.path.join(base_dir, log_filename)

    spider = TushareDocSpider(
        start_url='https://tushare.pro/document/2',
        output_file=full_excel_path,
        log_file=full_log_path,
        delay=1
    )
    spider.run()
# This Python file uses the following encoding: utf-8

# if __name__ == "__main__":
#     pass
