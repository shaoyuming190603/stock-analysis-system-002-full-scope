import os
import requests
from bs4 import BeautifulSoup
import time
import re
import logging
import random
from datetime import datetime
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed


class TushareDocSpider:
    """
    // class 版本：1.2.1 (第九工作表“接口名称(链接)”使用页面大标题)

    所属文件名：Tab01_API_tuShare_接口参数爬虫.py

    Tushare Pro 文档爬虫（多线程增强版）：提取接口名称、页面大标题、输入/输出参数的详细信息，
    并生成包含八个原始工作表及新增第九个“附加信息”工作表的 Excel 文件。
    附加信息包含：大类、小类、接口名称(链接)、接口、积分/权限、频率限制/分钟、限量/次、以及接口下方前五个段落全文。
    """
    def __init__(self, start_url, output_file='tushare_api_docs.xlsx', log_file='spider.log', delay=1, max_workers=5):
        self.start_url = start_url
        self.output_file = output_file
        self.delay = delay
        self.max_workers = max_workers
        self.log_file = log_file

        self._setup_logging()

        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })

        self.input_data = []
        self.output_data = []
        self.extra_data = []   # 附加信息列表

    def _setup_logging(self):
        self.logger = logging.getLogger('TushareSpider')
        self.logger.setLevel(logging.INFO)
        if self.logger.handlers:
            self.logger.handlers.clear()
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        log_dir = os.path.dirname(self.log_file)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir)
        file_handler = logging.FileHandler(self.log_file, encoding='utf-8')
        file_handler.setFormatter(formatter)
        self.logger.addHandler(file_handler)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        self.logger.addHandler(console_handler)

    def get_all_doc_links(self):
        self.logger.info(f"正在从起始页 {self.start_url} 提取所有文档链接...")
        try:
            resp = self.session.get(self.start_url)
            resp.encoding = 'utf-8'
            soup = BeautifulSoup(resp.text, 'html.parser')
            links = set()
            for a in soup.find_all('a', href=True):
                href = a['href']
                if 'document/2?doc_id=' in href:
                    full_url = urljoin(self.start_url, href)
                    links.add(full_url)
            self.logger.info(f"共找到 {len(links)} 个接口文档链接。")
            return list(links)
        except Exception as e:
            self.logger.error(f"获取链接失败：{e}", exc_info=True)
            return []

    def extract_title_above_interface(self, soup):
        section_keywords = ['输入参数', '输出参数', '返回数据', '错误码', '参数说明', '注意事项', '更新记录']
        h1 = soup.find('h1')
        if h1:
            title = h1.get_text().strip()
            if title and not any(keyword in title for keyword in section_keywords):
                return title
        h2 = soup.find('h2')
        if h2:
            title = h2.get_text().strip()
            if title and not any(keyword in title for keyword in section_keywords):
                return title
        interface_str = soup.find(string=re.compile(r'接口[：:]'))
        if not interface_str:
            interface_str = soup.find(string=re.compile(r'接口名称[：:]'))
        if not interface_str:
            return ""
        tag = interface_str.parent
        title_tag = tag.find_previous(['h1', 'h2', 'h3'])
        if not title_tag:
            title_tag = tag.find_previous(class_=lambda x: x and 'title' in x.lower())
        if title_tag:
            return title_tag.get_text().strip()
        return ""

    def _extract_params_details(self, soup, table_title):
        details = []
        header = soup.find(lambda tag: tag.name in ['h3', 'h4', 'div', 'p'] and table_title in tag.get_text())
        if not header:
            return details
        table = header.find_next('table')
        if not table:
            return details
        rows = table.find_all('tr')
        if len(rows) < 2:
            return details
        header_row = rows[0]
        header_cells = header_row.find_all(['th', 'td'])
        header_texts = [cell.get_text().strip().lower() for cell in header_cells]

        field_keywords = {
            'name': ['名称', '参数名', '字段', 'name'],
            'type': ['类型', 'type'],
            'desc': ['描述', '说明', 'desc'],
            'display': ['默认', '显示', '示例', 'default', 'display']
        }

        col_to_field = {}
        for idx, text in enumerate(header_texts):
            for field, keywords in field_keywords.items():
                if any(keyword in text for keyword in keywords):
                    col_to_field[idx] = field
                    break

        if not col_to_field:
            for idx in range(len(header_cells)):
                if idx == 0:
                    col_to_field[idx] = 'name'
                elif idx == 1:
                    col_to_field[idx] = 'type'
                elif idx == 2:
                    col_to_field[idx] = 'desc'
                elif idx == 3:
                    col_to_field[idx] = 'display'

        for row in rows[1:]:
            cells = row.find_all('td')
            if not cells:
                continue
            param = {'name': '', 'type': '', 'desc': '', 'display': ''}
            for idx, cell in enumerate(cells):
                if idx in col_to_field:
                    field = col_to_field[idx]
                    param[field] = cell.get_text().strip()
            if param['name']:
                details.append(param)
        return details

    def _extract_category_info(self, soup):
        """
        从左侧导航树中提取当前页面所属的大类和小类
        通过定位激活的 li，然后向上查找两级 li 中的 <a> 标签获取准确文本
        """
        main_cat = ''
        sub_cat = ''
        nav = soup.select_one('#jstree')
        if not nav:
            return main_cat, sub_cat

        active_li = nav.select_one('li.active, li.jstree-clicked')
        if not active_li:
            return main_cat, sub_cat

        # 当前 li 的父 ul（第三级菜单的容器）
        ul_level3 = active_li.find_parent('ul')
        if not ul_level3:
            return main_cat, sub_cat

        # 该 ul 的父 li（第二级，小类）
        li_level2 = ul_level3.find_parent('li')
        if li_level2:
            a_level2 = li_level2.find('a')
            if a_level2:
                sub_cat = a_level2.get_text(strip=True)
            # 再向上获取大类
            ul_level2 = li_level2.find_parent('ul')
            if ul_level2:
                li_level1 = ul_level2.find_parent('li')
                if li_level1:
                    a_level1 = li_level1.find('a')
                    if a_level1:
                        main_cat = a_level1.get_text(strip=True)
        return main_cat, sub_cat

    def _extract_integral_and_limit(self, soup):
        text = soup.get_text()
        integral_match = re.search(r'积分[^\d]*(\d+)', text)
        integral = integral_match.group(1) if integral_match else ''
        freq_match = re.search(r'(?:每分钟|调取)[^\d]*(\d+)\s*次', text)
        frequency = freq_match.group(1) if freq_match else ''
        limit_match = re.search(r'每次[^\d]*(\d+)\s*(?:条数据|行数据)', text)
        if not limit_match:
            limit_match = re.search(r'(\d+)\s*(?:条数据|行数据)', text)
        limit = limit_match.group(1) if limit_match else ''
        return integral, frequency, limit

    def _extract_first_five_paragraphs(self, soup, title_above):
        paragraphs = [''] * 5
        title_tag = None
        if title_above:
            title_tag = soup.find(['h1', 'h2'], string=re.compile(re.escape(title_above)))
        if not title_tag:
            title_tag = soup.find('h2')
        if not title_tag:
            return paragraphs
        p_tags = title_tag.find_next_siblings('p')
        for i, p in enumerate(p_tags[:5]):
            paragraphs[i] = p.get_text(strip=True)
        return paragraphs

    def parse_api_page(self, url):
        time.sleep(random.uniform(0.5, self.delay))
        self.logger.info(f"正在解析：{url}")

        session = requests.Session()
        session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })

        try:
            resp = session.get(url)
            resp.encoding = 'utf-8'
            soup = BeautifulSoup(resp.text, 'html.parser')
            text = soup.get_text()

            # 提取接口名称
            api_name = None
            interface_pattern = re.compile(r'(?:接口|接口名称)[：:]\s*([a-zA-Z0-9_]+)')
            match = interface_pattern.search(text)
            if match:
                api_name = match.group(1)
            else:
                for tag in soup.find_all(['h1', 'h2', 'h3', 'p', 'strong']):
                    if '接口' in tag.get_text():
                        text_line = tag.get_text()
                        parts = re.split(r'[：:]', text_line)
                        if len(parts) > 1:
                            candidate = parts[1].strip().split()[0]
                            if candidate and re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', candidate):
                                api_name = candidate
                                break

            if not api_name:
                self.logger.warning(f"无法从 {url} 提取接口名称，标记为失败。")
                return ('failed', None, None)

            if api_name == 'index_classify':
                self.logger.info(f"接口名称为 '{api_name}'，根据要求跳过此页面。")
                return ('skipped', None, None)

            title_above = self.extract_title_above_interface(soup)
            if not title_above:
                self.logger.warning(f"页面 {url} 未找到合适的标题，将使用空字符串。")
                title_above = ""  # 确保有值

            input_details = self._extract_params_details(soup, '输入参数')
            output_details = self._extract_params_details(soup, '输出参数')

            # 附加信息
            main_cat, sub_cat = self._extract_category_info(soup)
            # 确保大类小类不为空
            if not main_cat:
                main_cat = "未知大类"
            if not sub_cat:
                sub_cat = "未知小类"

            integral, frequency, limit = self._extract_integral_and_limit(soup)
            paras = self._extract_first_five_paragraphs(soup, title_above)

            extra_info = {
                'main_category': main_cat,
                'sub_category': sub_cat,
                'api_name': api_name,
                'title_above': title_above,          # 新增：页面大标题（用于第三列）
                'url': url,                           # 保存URL用于超链接
                'integral': integral,
                'frequency': frequency,
                'limit': limit,
                'para1': paras[0],
                'para2': paras[1],
                'para3': paras[2],
                'para4': paras[3],
                'para5': paras[4]
            }

            self.logger.info(f"成功解析接口：{api_name}，标题：{title_above}，输入参数数：{len(input_details)}，输出参数数：{len(output_details)}")
            return ('success', (api_name, title_above, url, input_details, output_details), extra_info)

        except Exception as e:
            self.logger.error(f"解析页面 {url} 时出错：{e}", exc_info=True)
            return ('failed', None, None)

    def run(self):
        self.logger.info("="*50)
        self.logger.info("爬虫程序开始运行（多线程模式）")
        self.logger.info(f"目标起始页：{self.start_url}")
        self.logger.info(f"输出文件：{self.output_file}")
        self.logger.info(f"日志文件：{self.log_file}")
        self.logger.info(f"线程池大小：{self.max_workers}")
        self.logger.info("="*50)

        doc_links = self.get_all_doc_links()
        if not doc_links:
            self.logger.error("未找到任何文档链接，程序退出。")
            return

        total_pages = len(doc_links)
        self.logger.info(f"开始遍历 {total_pages} 个接口页面...")

        input_local, output_local, extra_local = [], [], []
        success_count = skipped_count = failed_count = 0

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_url = {executor.submit(self.parse_api_page, link): link for link in doc_links}
            for future in as_completed(future_to_url):
                url = future_to_url[future]
                try:
                    status, data, extra = future.result()
                    if status == 'success':
                        api, title, url, inp, outp = data
                        input_local.append((api, title, url, inp))
                        output_local.append((api, title, url, outp))
                        extra_local.append(extra)
                        success_count += 1
                    elif status == 'skipped':
                        skipped_count += 1
                    else:
                        failed_count += 1
                except Exception as e:
                    self.logger.error(f"处理任务结果时发生异常（URL: {url}）：{e}", exc_info=True)
                    failed_count += 1

        self.input_data = input_local
        self.output_data = output_local
        self.extra_data = extra_local

        self.logger.info("="*50)
        self.logger.info("爬取完成，统计信息：")
        self.logger.info(f"总页面数：{total_pages}")
        self.logger.info(f"成功解析：{success_count}")
        self.logger.info(f"跳过页面：{skipped_count}")
        self.logger.info(f"失败页面：{failed_count}")
        self.logger.info("="*50)

        if self.input_data or self.output_data:
            self._save_to_excel()
        else:
            self.logger.warning("没有成功解析任何数据，Excel 文件未生成。")

    def _apply_styles(self, worksheet):
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row < 1 or max_col < 1:
            return
        dotted_side = Side(style='dotted')
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                left = dotted_side if c == 1 else None
                top = dotted_side if r == 1 else None
                right = dotted_side
                bottom = dotted_side
                border = Border(left=left, right=right, top=top, bottom=bottom)
                cell.border = border
                if cell.hyperlink:
                    font = Font(name='Arial', size=11, color="0000FF", underline="single")
                else:
                    font = Font(name='Arial', size=11)
                cell.font = font

    def _save_to_excel(self):
        self.logger.info("开始生成Excel文件...")

        output_dir = os.path.dirname(self.output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.logger.info(f"已创建目录：{output_dir}")

        max_params = 20
        columns = ['接口', '描述', '参数数'] + [f'参数{i+1}' for i in range(max_params)]

        sheets_config = [
            ('输入参数', self.input_data, 'name', True),
            ('输入类型', self.input_data, 'type', False),
            ('输入显示', self.input_data, 'display', False),
            ('输入描述', self.input_data, 'desc', False),
            ('输出参数', self.output_data, 'name', True),
            ('输出类型', self.output_data, 'type', False),
            ('输出显示', self.output_data, 'display', False),
            ('输出描述', self.output_data, 'desc', False)
        ]

        tab_colors = {
            '输入参数': 'FF0000', '输入类型': '9933FF', '输入显示': '9933FF', '输入描述': '9933FF',
            '输出参数': '0000FF', '输出类型': '00FF00', '输出显示': '00FF00', '输出描述': '00FF00'
        }

        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        def prepare_rows(data, attr_key, with_hyperlink):
            rows = []
            for api, title, url, details in data:
                values = [d.get(attr_key, '') for d in details]
                param_count = len(values)
                values_padded = values[:max_params] + [''] * (max_params - len(values))
                if with_hyperlink:
                    row = [api, (title, url), param_count] + values_padded
                else:
                    row = [api, title, param_count] + values_padded
                rows.append(row)
            return rows

        # 创建原始八个工作表
        for title, data_source, attr_key, hyperlink in sheets_config:
            ws = wb.create_sheet(title=title)
            ws.append(columns)
            rows = prepare_rows(data_source, attr_key, hyperlink)
            for row_data in rows:
                row_to_write = []
                for cell in row_data:
                    if isinstance(cell, tuple) and len(cell) == 2:
                        row_to_write.append(cell[0])
                    else:
                        row_to_write.append(cell)
                ws.append(row_to_write)
                if hyperlink:
                    last_row = ws.max_row
                    cell = ws.cell(row=last_row, column=2)
                    cell.hyperlink = row_data[1][1]
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            for i in range(4, 4 + max_params):
                ws.column_dimensions[get_column_letter(i)].width = 15
            self._apply_styles(ws)
            ws.sheet_properties.tabColor = tab_colors[title]

        # 创建第九个工作表：附加信息（第三列使用页面大标题，带超链接）
        if self.extra_data:
            ws = wb.create_sheet(title="附加信息")
            headers = ['大类', '小类', '接口名称（链接）', '接口', '积分/权限', '频率限制/分钟', '限量/次',
                       '段落1', '段落2', '段落3', '段落4', '段落5']
            ws.append(headers)

            for item in self.extra_data:
                row = [
                    item.get('main_category', ''),
                    item.get('sub_category', ''),
                    item.get('title_above', ''),       # 第三列：页面大标题（如“A股日线行情”）
                    item.get('api_name', ''),           # 第四列：纯文本接口名称
                    item.get('integral', ''),
                    item.get('frequency', ''),
                    item.get('limit', ''),
                    item.get('para1', ''),
                    item.get('para2', ''),
                    item.get('para3', ''),
                    item.get('para4', ''),
                    item.get('para5', '')
                ]
                ws.append(row)
                # 为第三列设置超链接
                last_row = ws.max_row
                cell = ws.cell(row=last_row, column=3)
                cell.hyperlink = item.get('url', '')

            # 设置列宽
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 30   # 带链接的接口名称列加宽
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            for i in range(8, 13):  # H到L列
                ws.column_dimensions[get_column_letter(i)].width = 40

            self._apply_styles(ws)
            ws.sheet_properties.tabColor = "FFA500"  # 橙色

        wb.save(self.output_file)
        self.logger.info(f"Excel文件已成功保存到：{self.output_file}")


if __name__ == "__main__":
    base_dir = r"C:\01_My_documents\01_pythonLearn\我的PNP_MCSD大项目\Tab01_数据设置\Tab0101_TuShareAPI接口参数"
    save_filename = "tuShare接口输入输出参数清单Ae9.xlsx"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"tuShare爬虫日志_{timestamp}.log"

    logging.basicConfig(filename=log_filename, level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')

    full_excel_path = os.path.join(base_dir, save_filename)
    full_log_path = os.path.join(base_dir, log_filename)

    spider = TushareDocSpider(
        start_url='https://tushare.pro/document/2',
        output_file=full_excel_path,
        log_file=full_log_path,
        delay=1,
        max_workers=5
    )
    spider.run()